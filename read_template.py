import openpyxl
import sys

file_path = r"c:\SQA\Bảng tính không có tiêu đề.xlsx"
with open(r"c:\SQA\template_info.txt", "w", encoding="utf-8") as f:
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        f.write(f"Sheets: {wb.sheetnames}\n")
        
        if 'Phan Anh ' in wb.sheetnames:
            ws = wb['Phan Anh ']
        else:
            ws = wb.active
            
        f.write(f"Using sheet: {ws.title}\n")
        
        # Read the first 15 rows
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
            f.write(f"Row {i}: {row}\n")
    except Exception as e:
        f.write(f"Error: {e}\n")
