import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Postman API Test"

HEADER_BG = "C6EFCE"
GROUP_BG  = "BDD7EE"
ROLLBACK_BG = "FFEB9C"

def setcell(ws, row, col, value, bold=False, bg=None, wrap=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=11)
    c.alignment = Alignment(wrap_text=wrap, vertical="top")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

# ---- Header info ----
ws.cell(1, 1, "Implementer:").font = Font(bold=True)
ws.cell(1, 2, "Phan Anh")
ws.cell(1, 8, "Project Link: https://github.com/NHPhanAnh/SQA-Test-Script.git")

ws.cell(2, 1, "Features:").font = Font(bold=True)
ws.cell(2, 2, "Authentication Verification, Token Validation Flow, Register Candidate")

# ---- Tools ----
ws.cell(4, 1, "1. Tools and Libraries:").font = Font(bold=True)
tools_text = (
    "- Postman v11 (API Testing Tool)\n"
    "- Newman (CLI runner for Postman collections - run: newman run collection.json)\n"
    "- MySQL CLI via docker exec (CheckDB & Rollback)"
)
ws.cell(5, 2, tools_text)
ws.row_dimensions[5].height = 60

# ---- Scope ----
ws.cell(7, 1, "2. Scope of Testing:").font = Font(bold=True)
setcell(ws, 8, 2, "Function / API", bold=True)
setcell(ws, 8, 5, "Reasons for NOT testing", bold=True)

scope_rows = [
    ("POST /auth/login", ""),
    ("POST /auth/logout", ""),
    ("POST /auth/register-candidate", ""),
    ("POST /auth/verify-token", ""),
    ("GET  /auth/request-resend-verify-email", ""),
    ("PUT  /auth/verify-email/{token}", ""),
    ("GET  /auth/request-reset-password", ""),
    ("PUT  /auth/reset-password", ""),
    ("POST /auth/register-org-admin",
     "Requires multipart/form-data with logo file upload. Complex binary attachment makes automated collection-level testing impractical. Covered by manual QA."),
    ("Google OAuth2 / Social Login",
     "External provider. Google is responsible for its own auth logic. We only mock its token response. Integration behavior verified via manual E2E test."),
]
for i, (fn, reason) in enumerate(scope_rows):
    ws.cell(9 + i, 2, fn)
    ws.cell(9 + i, 5, reason)

# ---- Test cases ----
r = 21
ws.cell(r, 1, "3. Test Cases").font = Font(bold=True, size=12)
r += 1

hdr = ["", "Test Case ID", "Test Objective", "", "Input", "Expected Output", "", "Notes"]
for ci, h in enumerate(hdr, 1):
    setcell(ws, r, ci, h, bold=True, bg=HEADER_BG if h else None)
r += 1

TC = [
    # (id, objective, input, expected, notes)  -- group rows have only id filled
    ("--- 1. Authentication Flow ---", "", "", "", ""),
    ("TC-POST-01", "Login with valid credentials",
     'username="sqapananh2026", password="Test@12345"',
     'HTTP 200; response.data.token exists; accessToken saved as collection variable',
     "Happy path"),
    ("TC-POST-02", "Login with wrong password",
     'username="sqapananh2026", password="WrongPassword999"',
     'HTTP != 200; response.message contains error description',
     "Invalid path"),
    ("TC-POST-03", "Login with non-existent username",
     'username="user_xyz_notexist", password="Test@12345"',
     'HTTP != 200; response.message contains user-not-found error',
     "Invalid path"),
    ("TC-POST-04", "Login with empty username",
     'username="", password="Test@12345"',
     'HTTP 400 Bad Request; validation error returned',
     "Edge case"),
    ("TC-POST-05", "Login with empty password",
     'username="sqapananh2026", password=""',
     'HTTP 400 Bad Request; validation error returned',
     "Edge case"),
    ("TC-POST-06", "Verify valid access token",
     'Authorization: Bearer {{accessToken}} (captured from TC-POST-01)',
     'HTTP 200; response.data.isValid === true',
     "Happy path"),
    ("TC-POST-07", "Verify token - no Authorization header sent",
     'No Authorization header',
     'HTTP 200; response.data.isValid === false',
     "Invalid path"),
    ("TC-POST-08", "Logout with valid token",
     'Authorization: Bearer {{accessToken}}',
     'HTTP 200; logout successful',
     "Happy path"),
    ("--- 2. Register Candidate ---", "", "", "", ""),
    ("TC-POST-09", "Register candidate with fully valid data",
     'username="sqapananh2027", password="Test@12345", fullName="SQA Test User", email="sqa_2027@mailinator.com"',
     'HTTP 200; user created. CheckDB: SELECT * FROM users WHERE username="sqapananh2027" => 1 row',
     "Happy path + CheckDB"),
    ("TC-POST-10", "Register with duplicate username",
     'username="sqapananh2027" (same as TC-POST-09)',
     'HTTP != 200; duplicate username error in response',
     "Invalid path"),
    ("TC-POST-11", "Register with invalid email format",
     'email="not-an-email"',
     'HTTP 400; email format validation error returned',
     "Edge case"),
    ("TC-POST-12", "Register with empty fullName",
     'fullName=""',
     'HTTP 400; fullName required error returned',
     "Edge case"),
    ("TC-POST-13 [ROLLBACK]", "Delete test user to restore DB state",
     'docker exec mysql-cvconnect mysql -uroot -p123456789\n=> DELETE FROM `cvconnect-user-service`.users WHERE username="sqapananh2026" OR username="sqapananh2027";',
     'CheckDB: SELECT COUNT(*) WHERE username IN ("sqapananh2026", "sqapananh2027") => 0 rows',
     "Rollback"),
    ("--- 3. Email Verification Token ---", "", "", "", ""),
    ("TC-POST-14", "Request resend verify email - valid identifier",
     'identifier="sqa_2027@mailinator.com"',
     'HTTP 200; resend email request accepted',
     "Happy path"),
    ("TC-POST-15", "Request resend verify email - non-existent identifier",
     'identifier="does_not_exist_xyz@test.com"',
     'HTTP != 200; user not found error',
     "Invalid path"),
    ("TC-POST-16", "Verify email - invalid token format",
     'token="invalid-token-xyz-12345" (path variable)',
     'HTTP != 200; invalid or expired token error',
     "Invalid path"),
    ("--- 4. Reset Password Token ---", "", "", "", ""),
    ("TC-POST-17", "Request reset password - valid identifier",
     'identifier="sqa_test_pananh@mailinator.com"',
     'HTTP 200; reset password email queued successfully',
     "Happy path"),
    ("TC-POST-18", "Request reset password - non-existent identifier",
     'identifier="no_such_user_xyzabc"',
     'HTTP != 200; user not found error',
     "Invalid path"),
    ("TC-POST-19", "Reset password - invalid token",
     'token="this-is-an-invalid-token", newPassword="NewPass@123"',
     'HTTP != 200; invalid token error returned',
     "Invalid path"),
    ("TC-POST-20", "Reset password - empty token field",
     'token="", newPassword="NewPass@123"',
     'HTTP 400; token required validation error',
     "Edge case"),
]

for row_data in TC:
    tc_id = row_data[0]
    is_group   = tc_id.startswith("---")
    is_rollback = "ROLLBACK" in tc_id

    bg = GROUP_BG if is_group else (ROLLBACK_BG if is_rollback else None)

    ws.cell(r, 1, "")
    for ci, val in enumerate(row_data, 2):
        setcell(ws, r, ci, val, bold=is_group, bg=bg)
    ws.row_dimensions[r].height = 55
    r += 1

# ---- Column widths ----
widths = [3, 22, 42, 3, 55, 58, 3, 18]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

out_path = r"C:\SQA\SQA-Test-Script\postman\Tool_Test_Postman_PhanAnh.xlsx"
wb.save(out_path)
print("SAVED:", out_path)
